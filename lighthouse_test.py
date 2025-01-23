import os
import subprocess
import sys
import signal
import time
import json
import shutil
from tqdm import tqdm
from datetime import datetime
from termcolor import colored
import threading
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# Fonction pour effacer le terminal
def clear_terminal():
    os.system('cls' if os.name == 'nt' else 'clear')

# Variable globale pour gérer l'état d'arrêt
stop_signal = False

# Fonction pour arrêter proprement le script lors d'un KeyboardInterrupt
def signal_handler(sig, frame):
    global stop_signal
    stop_signal = True  # Mettre le drapeau à True pour indiquer l'arrêt
    print(colored("\nMerci d'avoir utilisé LightHouse Test", "white"))  # Message à afficher
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)

# Fonction pour afficher le logo
def display_logo():
    logo = """
 _       _         _      _    _    _                            _______             _
| |     (_)       | |    | |  | |  | |                          |__   __|           | |
| |      _   __ _ | |__  | |_ | |__| |  ___   _   _  ___   ___     | |     ___  ___ | |_
| |     | | / _` || '_ \ | __||  __  | / _ \ | | | |/ __| / _ \    | |    / _ \/ __|| __|
| |____ | || (_| || | | || |_ | |  | || (_) || |_| |\__ \|  __/    | |   |  __/\__ \| |_
|______||_| \__, ||_| |_| \__||_|  |_| \___/  \__,_||___/ \___|    |_|    \___||___/ \__|
             __/ |
            |___/                                                      by BreakingTech.fr
    """
    print(colored(logo, "red"))

# Menu principal
def menu():
    print(colored("\nBienvenue dans Lighthouse Test", "cyan"))
    print("1. Tester une URL commençant par http, https ou www.")
    print("2. Glissez-déposez un fichier .txt contenant une liste d'URLs")
    print("3. Quitter le programme")
    choice = input(colored("\nFaites votre choix (1, 2 ou 3): ", "yellow")).strip()
    return choice

def run_tests(urls):
    data_results = []
    
    for url in urls:
        # Analyse mobile
        mobile_data = get_lighthouse_scores(url, "mobile")
        if mobile_data is None:
            continue

        # Analyse desktop
        desktop_data = get_lighthouse_scores(url, "desktop")
        if desktop_data is None:
            continue

        # Stockage des résultats pour cette URL uniquement
        current_result = {
            'url': url,
            'mobile_data': mobile_data,
            'desktop_data': desktop_data
        }

        # Ajout aux résultats globaux
        data_results.append(current_result)

        # Affichage du tableau uniquement pour l'URL courante
        display_scores([current_result])

    # Transformation de data_results en un dictionnaire pour l'exportation Excel
    data_dict = {
        'URL': [],
        'Stratégie': [],
        'FCP (s)': [],
        'LCP (s)': [],
        'TBT (ms)': [],
        'CLS': [],
        'Speed Index (s)': [],
        'Performance Score': [],
        'Accessibility Score': [],
        'Best Practices Score': [],
        'SEO Score': []
    }

    for result in data_results:
        url = result['url']
        mobile_data = json.loads(result['mobile_data'])
        desktop_data = json.loads(result['desktop_data'])

        # Extraction des scores mobile
        mobile_scores = mobile_data['categories']
        mobile_audits = mobile_data['audits']
        data_dict['URL'].append(url)
        data_dict['Stratégie'].append("Mobile")
        data_dict['FCP (s)'].append(mobile_audits['first-contentful-paint']['numericValue'] / 1000)
        data_dict['LCP (s)'].append(mobile_audits['largest-contentful-paint']['numericValue'] / 1000)
        data_dict['TBT (ms)'].append(mobile_audits['total-blocking-time']['numericValue'])
        data_dict['CLS'].append(mobile_audits['cumulative-layout-shift']['numericValue'])
        data_dict['Speed Index (s)'].append(mobile_audits['speed-index']['numericValue'] / 1000)
        data_dict['Performance Score'].append(int(mobile_scores['performance']['score'] * 100))
        data_dict['Accessibility Score'].append(int(mobile_scores['accessibility']['score'] * 100))
        data_dict['Best Practices Score'].append(int(mobile_scores['best-practices']['score'] * 100))
        data_dict['SEO Score'].append(int(mobile_scores['seo']['score'] * 100))

        # Extraction des scores desktop
        desktop_scores = desktop_data['categories']
        desktop_audits = desktop_data['audits']
        data_dict['URL'].append(url)
        data_dict['Stratégie'].append("Desktop")
        data_dict['FCP (s)'].append(desktop_audits['first-contentful-paint']['numericValue'] / 1000)
        data_dict['LCP (s)'].append(desktop_audits['largest-contentful-paint']['numericValue'] / 1000)
        data_dict['TBT (ms)'].append(desktop_audits['total-blocking-time']['numericValue'])
        data_dict['CLS'].append(desktop_audits['cumulative-layout-shift']['numericValue'])
        data_dict['Speed Index (s)'].append(desktop_audits['speed-index']['numericValue'] / 1000)
        data_dict['Performance Score'].append(int(desktop_scores['performance']['score'] * 100))
        data_dict['Accessibility Score'].append(int(desktop_scores['accessibility']['score'] * 100))
        data_dict['Best Practices Score'].append(int(desktop_scores['best-practices']['score'] * 100))
        data_dict['SEO Score'].append(int(desktop_scores['seo']['score'] * 100))

    # Exportation unique vers Excel avec le dictionnaire final
    export_to_excel(data_dict)

        
# Fonction pour tester une seule URL
def test_single_url():
    url = input(colored("Veuillez entrer l'URL (http, https ou www): ", "yellow")).strip()  # Changement en jaune
    if not (url.startswith(('http://', 'https://', 'www.'))):
        print(colored("L'URL doit commencer par http, https, ou www.", "red"))
        sys.exit(1)
    
    if url.startswith('www.'):
        url = 'http://' + url
    
    return [url]

# Fonction pour tester un fichier .txt contenant des URLs
def test_urls_from_file():
    print(colored("Glissez-déposez ici un fichier .txt contenant les URL :", "yellow"))  # Changement en jaune
    file_path = input().strip()

    # Remplacer les séquences \ avant des espaces par de simples espaces
    file_path = file_path.replace("\\ ", " ")

    if not os.path.isfile(file_path):
        print(colored("Le fichier spécifié n'existe pas.", "red"))
        sys.exit(1)

    with open(file_path, 'r') as file:
        urls = [line.strip() for line in file if line.strip().startswith(('http://', 'https://', 'www.'))]
    
    if not urls:
        print(colored("Aucune URL valide trouvée dans le fichier.", "red"))
        sys.exit(1)
    
    for i in range(len(urls)):
        if urls[i].startswith('www.'):
            urls[i] = 'http://' + urls[i]

    return urls

# Fonction pour exécuter Lighthouse et récupérer les scores pour une URL
def get_lighthouse_scores(url, strategy):
    print(colored(f"\nAnalyse de {url} ({strategy})", "magenta"))

    # Construire dynamiquement le chemin de l'exécutable Lighthouse
    user_profile = os.getenv('USERPROFILE')
    lighthouse_path = os.path.join(user_profile, r'AppData\Roaming\npm\lighthouse.cmd')

    if strategy == "mobile":
        command = [
            lighthouse_path, url, 
            '--only-categories=performance,accessibility,best-practices,seo', 
            '--output=json', '--emulated-form-factor=mobile', '--quiet', '--chrome-flags=--headless'
        ]
    else:
        command = [
            lighthouse_path, url, 
            '--only-categories=performance,accessibility,best-practices,seo', 
            '--output=json', '--preset=desktop', '--quiet', '--chrome-flags=--headless'
        ]

    progress_bar = tqdm(total=100, desc="Analyse en cours", bar_format="{l_bar}{bar} [Temps écoulé: {elapsed}, Temps restant: {remaining}]")

    def update_progress_bar():
        while progress_bar.n < 100 and not stop_signal:
            time.sleep(0.5)
            progress_bar.update(1)

    thread = threading.Thread(target=update_progress_bar)
    thread.start()

    result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, encoding='utf-8')

    progress_bar.n = 100
    progress_bar.close()

    thread.join()  # Assurez-vous que le thread se termine avant de continuer

    if result.returncode != 0:
        print(colored(f"Erreur lors de l'analyse de {url}", "red"))
        print(f"STDOUT : {result.stdout}")
        print(f"STDERR : {result.stderr}")
        return None
    return result.stdout  # Retourne la sortie JSON

# Fonction pour extraire et afficher les scores sous forme de tableau
def display_scores(data_results):
    import json
    try:
        for result in data_results:
            url = result['url']
            mobile_data = result['mobile_data']
            desktop_data = result['desktop_data']

            # Scores mobile
            json_data_mobile = json.loads(mobile_data)
            categories_mobile = json_data_mobile.get('categories', {})
            scores_mobile = {
                "Performance": int(categories_mobile.get('performance', {}).get('score', 0) * 100),
                "Accessibility": int(categories_mobile.get('accessibility', {}).get('score', 0) * 100),
                "Best Practices": int(categories_mobile.get('best-practices', {}).get('score', 0) * 100),
                "SEO": int(categories_mobile.get('seo', {}).get('score', 0) * 100)
            }

            # Extraire les scores de performance mobile
            audits_mobile = json_data_mobile['audits']
            fcp_mobile = audits_mobile['first-contentful-paint']['numericValue'] / 1000
            lcp_mobile = audits_mobile['largest-contentful-paint']['numericValue'] / 1000
            tbt_mobile = audits_mobile['total-blocking-time']['numericValue']
            cls_mobile = float(audits_mobile['cumulative-layout-shift']['numericValue'])
            speed_index_mobile = audits_mobile['speed-index']['numericValue'] / 1000

            # Scores desktop
            json_data_desktop = json.loads(desktop_data)
            categories_desktop = json_data_desktop.get('categories', {})
            scores_desktop = {
                "Performance": int(categories_desktop.get('performance', {}).get('score', 0) * 100),
                "Accessibility": int(categories_desktop.get('accessibility', {}).get('score', 0) * 100),
                "Best Practices": int(categories_desktop.get('best-practices', {}).get('score', 0) * 100),
                "SEO": int(categories_desktop.get('seo', {}).get('score', 0) * 100)
            }

            # Extraire les scores de performance desktop
            audits_desktop = json_data_desktop['audits']
            fcp_desktop = audits_desktop['first-contentful-paint']['numericValue'] / 1000
            lcp_desktop = audits_desktop['largest-contentful-paint']['numericValue'] / 1000
            tbt_desktop = audits_desktop['total-blocking-time']['numericValue']
            cls_desktop = float(audits_desktop['cumulative-layout-shift']['numericValue'])
            speed_index_desktop = audits_desktop['speed-index']['numericValue'] / 1000

            # Fonction de formatage dynamique pour CLS
            def format_cls(cls_value):
                # Vérifier si CLS est un nombre entier
                if cls_value.is_integer():
                    return str(int(cls_value))  # Afficher sans décimales pour les entiers
                # Limiter à 5 chiffres après la virgule pour les valeurs décimales
                return f"{cls_value:.6f}".rstrip('0').rstrip('.')  # Supprimer les zéros et le point final inutiles

            # Affichage du tableau des scores pour chaque URL
            print(colored("+" + "-"*81 + "+", "cyan"))
            print(colored("| {:<31} | {:^21} | {:^21} |".format("Critères", "Scores Mobile", "Scores Desktop"), "cyan"))
            print(colored("+" + "-"*81 + "+", "cyan"))

            # Afficher tous les scores combinés dans un seul tableau avec les scores centrés
            print(f"| {colored('Performance', 'white'):<40} | {colored(scores_mobile['Performance'], 'green' if scores_mobile['Performance'] >= 90 else 'yellow' if scores_mobile['Performance'] >= 50 else 'red'):^30} | {colored(scores_desktop['Performance'], 'green' if scores_desktop['Performance'] >= 90 else 'yellow' if scores_desktop['Performance'] >= 50 else 'red'):^30} |")
            print(f"| {colored('Accessibility', 'white'):<40} | {colored(scores_mobile['Accessibility'], 'green' if scores_mobile['Accessibility'] >= 90 else 'yellow' if scores_mobile['Accessibility'] >= 50 else 'red'):^30} | {colored(scores_desktop['Accessibility'], 'green' if scores_desktop['Accessibility'] >= 90 else 'yellow' if scores_desktop['Accessibility'] >= 50 else 'red'):^30} |")
            print(f"| {colored('Best Practices', 'white'):<40} | {colored(scores_mobile['Best Practices'], 'green' if scores_mobile['Best Practices'] >= 90 else 'yellow' if scores_mobile['Best Practices'] >= 50 else 'red'):^30} | {colored(scores_desktop['Best Practices'], 'green' if scores_desktop['Best Practices'] >= 90 else 'yellow' if scores_desktop['Best Practices'] >= 50 else 'red'):^30} |")
            print(f"| {colored('SEO', 'white'):<40} | {colored(scores_mobile['SEO'], 'green' if scores_mobile['SEO'] >= 90 else 'yellow' if scores_mobile['SEO'] >= 50 else 'red'):^30} | {colored(scores_desktop['SEO'], 'green' if scores_desktop['SEO'] >= 90 else 'yellow' if scores_desktop['SEO'] >= 50 else 'red'):^30} |")
            print(colored("+" + "-"*81 + "+", "cyan"))

            # Afficher les scores de performance combinés pour mobile et desktop avec scores centrés
            print(f"| {colored('First Contentful Paint (FCP)', 'white'):<40} | {colored(f'{fcp_mobile:.1f}s' if fcp_mobile % 1 else f'{int(fcp_mobile)}s', 'green' if fcp_mobile <= 1.82 else 'yellow' if fcp_mobile <= 3.01 else 'red'):^30} | {colored(f'{fcp_desktop:.1f}s' if fcp_desktop % 1 else f'{int(fcp_desktop)}s', 'green' if fcp_desktop <= 2.31 else 'red'):^30} |")
            print(f"| {colored('Largest Contentful Paint (LCP)', 'white'):<40} | {colored(f'{lcp_mobile:.1f}s' if lcp_mobile % 1 else f'{int(lcp_mobile)}s', 'green' if lcp_mobile <= 2.52 else 'yellow' if lcp_mobile <= 4.01 else 'red'):^30} | {colored(f'{lcp_desktop:.1f}s' if lcp_desktop % 1 else f'{int(lcp_desktop)}s', 'green' if lcp_desktop <= 1.21 else 'yellow' if lcp_desktop <= 2.41 else 'red'):^30} |")
            print(f"| {colored('Total Blocking Time (TBT)', 'white'):<40} | {colored(f'{tbt_mobile:.1f}ms' if tbt_mobile % 1 else f'{int(tbt_mobile)}ms', 'green' if tbt_mobile <= 200 else 'yellow' if tbt_mobile <= 600 else 'red'):^30} | {colored(f'{tbt_desktop:.1f}ms' if tbt_desktop % 1 else f'{int(tbt_desktop)}ms', 'green' if tbt_desktop <= 150 else 'yellow' if tbt_desktop <= 350 else 'red'):^30} |")
            print(f"| {colored('Cumulative Layout Shift (CLS)', 'white'):<40} | {colored(format_cls(cls_mobile), 'green' if cls_mobile <= 0.1 else 'orange' if cls_mobile <= 0.25 else 'red'):^30} | {colored(format_cls(cls_desktop), 'green' if cls_desktop <= 0.1 else 'orange' if cls_desktop <= 0.25 else 'red'):^30} |")
            print(f"| {colored('Speed Index', 'white'):<40} | {colored(f'{speed_index_mobile:.1f}s' if speed_index_mobile % 1 else f'{int(speed_index_mobile)}s', 'green' if speed_index_mobile <= 3.42 else 'yellow' if speed_index_mobile <= 5.82 else 'red'):^30} | {colored(f'{speed_index_desktop:.1f}s' if speed_index_desktop % 1 else f'{int(speed_index_desktop)}s', 'green' if speed_index_desktop <= 1.32 else 'yellow' if speed_index_desktop <= 2.31 else 'red'):^30} |")
            print(colored("+" + "-"*81 + "+", "cyan"))

    except json.JSONDecodeError:
            print(colored("Erreur lors de l'analyse des données de Lighthouse.", "red"))

# Fonction pour définir la couleur de remplissage
def set_fill_color(header, score, device_type):
    # Convertir en float si score est une chaîne de caractères
    try:
        score = float(score)
    except ValueError:
        # Si la conversion échoue, laisser score tel quel
        return None

    # Normalisation des valeurs flottantes avec un arrondi à une décimale
    score = round(score, 1)

    if header in ["FCP (s)", "LCP (s)", "TBT (ms)", "Speed Index (s)", "CLS"]:
        if device_type == "mobile":
            if header == 'FCP (s)':
                if score <= 1.82:
                    return '5ED050'  # Vert
                elif score <= 3.01:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'LCP (s)':
                if score <= 2.52:
                    return '5ED050'  # Vert
                elif score <= 4.01:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'TBT (ms)':
                if score <= 200:
                    return '5ED050'  # Vert
                elif score <= 600:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'Speed Index (s)':
                if score <= 3.42:
                    return '5ED050'  # Vert
                elif score <= 5.82:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'CLS':
                if score <= 0.1:
                    return '5ED050'  # Vert
                elif score <= 0.25:
                    return 'FFA500'  # Orange
                else:
                    return 'FF0000'  # Rouge
        elif device_type == "desktop":
            if header == 'FCP (s)':
                if score <= 0.94:
                    return '5ED050'  # Vert
                elif score <= 1.6:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'LCP (s)':
                if score <= 1.21:
                    return '5ED050'  # Vert
                elif score <= 2.41:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'TBT (ms)':
                if score <= 150:
                    return '5ED050'  # Vert
                elif score <= 350:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'Speed Index (s)':
                if score <= 1.32:
                    return '5ED050'  # Vert
                elif score <= 2.31:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
            elif header == 'CLS':
                if score <= 0.1:
                    return '5ED050'  # Vert
                elif score <= 0.25:
                    return 'FFF055'  # Jaune
                else:
                    return 'FF0000'  # Rouge
    return None

def export_to_excel(data):
    # Création d'un nouveau classeur Excel
    workbook = Workbook()
    sheet = workbook.active

    # Réorganisation des en-têtes dans l'ordre spécifié
    headers = [
        'URL', 'Stratégie', 'Performance Score', 'Accessibility Score', 'Best Practices Score', 'SEO Score',
        'FCP (s)', 'LCP (s)', 'TBT (ms)', 'CLS', 'Speed Index (s)'
    ]
    sheet.append(headers)

    # Appliquer la mise en forme en gras, l'alignement centré et le contour noir aux en-têtes
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)  # Mettre en gras
        cell.alignment = Alignment(horizontal='center')  # Centrer l'alignement
        cell.border = Border(left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'))  # Contour noir

        # Appliquer la couleur de fond pour la première ligne
        if cell.value is not None:
            fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type="solid")
            cell.fill = fill

    # Remplissage des données en suivant le nouvel ordre des en-têtes
    for i in range(len(data['URL'])):
        row = [data[header][i] for header in headers]
        sheet.append(row)

        # Récupérer device_type en fonction de la colonne "Stratégie"
        device_type = "mobile" if data["Stratégie"][i] == "Mobile" else "desktop"

        # Appliquer la couleur de fond et l'alignement selon les scores
        for col_idx, header in enumerate(headers):
            score = data[header][i]
            cell = sheet.cell(row=i + 2, column=col_idx + 1)

            # Formater selon les exigences pour les colonnes spécifiques
            if header in ["FCP (s)", "LCP (s)", "TBT (ms)", "Speed Index (s)"]:
                cell.number_format = '0.0' if isinstance(score, float) and score % 1 != 0 else '0'
            elif header == "CLS":
                cell.number_format = '0.00000' if isinstance(score, float) and score % 1 != 0 else '0'

            # Appliquer la couleur de fond selon les scores et device_type
            fill_color = None
            if "Score" in header:
                if score >= 90:
                    fill_color = '5ED050'  # Vert
                elif score >= 50:
                    fill_color = 'FFF055'  # Jaune
                else:
                    fill_color = 'FF0000'  # Rouge
            else:
                fill_color = set_fill_color(header, score, device_type)

            # Appliquer le remplissage
            if fill_color:
                fill = PatternFill(start_color='FF' + fill_color, end_color='FF' + fill_color, fill_type="solid")
                cell.fill = fill

            # Appliquer l'alignement
            alignment = Alignment(horizontal='center')
            if header != "URL":
                cell.alignment = alignment

            # Appliquer le contour noir si la cellule contient du contenu
            if cell.value is not None:
                cell.border = Border(left=Side(style='thin', color='000000'),
                                     right=Side(style='thin', color='000000'),
                                     top=Side(style='thin', color='000000'),
                                     bottom=Side(style='thin', color='000000'))

    # Ajuster la largeur des colonnes
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Obtenir la lettre de la colonne
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # Ajouter un peu d'espace
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Sauvegarde du fichier Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'rapport_performance_{timestamp}.xlsx'
    script_directory = os.path.dirname(os.path.abspath(__file__))
    full_path = os.path.join(script_directory, filename)
    workbook.save(full_path)
    print(colored(f"\nLes scores ont été exportés vers '{filename}'", "green"))


# Fonction principale du programme
def main():
    clear_terminal()
    display_logo()

    while True:
        choice = menu()
        if choice == "1":
            urls = test_single_url()
        elif choice == "2":
            urls = test_urls_from_file()
        elif choice == "3":
            print(colored("Merci d'avoir utilisé LightHouse Test", "white"))
            sys.exit(0)
        else:
            print(colored("Choix invalide. Veuillez réessayer.", "red"))
            continue
        
        # Appel de la fonction run_tests pour analyser les URLs
        run_tests(urls)

if __name__ == "__main__":
    main()